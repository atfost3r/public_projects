#! python3

# snedDuesReminders.py - Sends emails based on payment status

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")

lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log into email account
smtpObj = smtplib.SMTP("stmp.gmail.com", 587)
smtpObj.ehlo()
smtpObj.starttls()
