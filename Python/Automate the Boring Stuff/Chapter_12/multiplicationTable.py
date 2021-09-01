#! python3
# multiplicationTable.py - The program will take a command line numerical input, N, and create an NxN matrix in an excel file

import openpyxl, sys
from openpyxl.styles import Font, Style

# Get the input from the commandline

N = int(sys.argv[1])

# Open Workbook
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

# create the font objects that make up the outter part of the table and the inner part
fontObj = Font(name="Times New Roman", bold="True")
styleObj = Style(font=fontObj)

# Create array to right out the square
for i in range(1, N + 1):
    for j in range(1, N + 1):
        sheet.cell(row=i + 1, column=j + 1).value = i * j
# TODO: Write out the "header" portions
for i in range(2, N + 2):
    sheet.cell(row=i, column=1).value = i - 1
    sheet.cell(row=1, column=i).value = i - 1


# Save and close workbook
wb.save("multiplicationTable.xlsx")
